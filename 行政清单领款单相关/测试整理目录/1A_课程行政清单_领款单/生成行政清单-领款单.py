# ============================================================
# 导入库模块
# ============================================================

# 标准库
import os
import re
import datetime
import sys
from collections import Counter

# 第三方库
import pandas as pd
from docxtpl import DocxTemplate
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# ============================================================
# 路径配置（已根据用户文件结构集中整理）
# ============================================================

# 把配置文件所在目录加入系统路径
sys.path.append(
    os.path.join(
        os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")),
        "0_模板文件及初始化",
    )
)


# 🔧 引用全局配置文件
from config_paths import (
    course_file,
    receipt_file,
    template_path,
    template_lingkuan_path,
    output_dir,
    excel_path,
    template_output_spending_actual_path,
)


# ============================================================
# 以下为业务处理逻辑
# ============================================================

try:
    df_course = pd.read_csv(course_file, encoding="utf-8-sig")
except Exception as e:
    print("❌ 课程数据读取失败，请检查文件路径或编码。")
    raise e

if df_course.empty:
    raise ValueError("❌ 课程数据为空，无法生成文档。")

row_index = 1 if len(df_course) > 1 else 0
row = df_course.iloc[row_index]

# ============================================================
# 读取收据数据（含自动容错）
# ============================================================
try:
    df_receipt = pd.read_csv(receipt_file, encoding="utf-8-sig")
except Exception as e:
    print("❌ 收据数据读取失败，请检查文件路径或编码。")
    raise e

if df_receipt.empty:
    print("⚠️ 收据数据为空，将跳过统计逻辑。")
    df_receipt = pd.DataFrame(columns=["會員類別", "扣減", "總額"])

# ============================================================
# 一、课程名称与编号
# ============================================================
课程名称 = str(row["名稱"]).strip()
课程编号 = str(row["編號"]).strip()
课程名称_cleaned = 课程名称.replace(课程编号, "")
课程名称_cleaned = re.split(r"\|", 课程名称_cleaned)[0].strip()

# ============================================================
# 二、收费拆分与清洗
# ============================================================
收费原文 = str(row["收費"]).strip()
parts = 收费原文.split("|")
费用_会员_raw = parts[0] if len(parts) >= 1 else ""
费用_非会员_raw = parts[1] if len(parts) >= 2 else ""


def clean_fee(fee_string):
    fee = re.split(r"\(", fee_string)[0].strip()
    fee = re.sub(r"\.00$", "", fee)
    fee = re.sub(r"(\.\d)0$", r"\1", fee)
    return fee


费用_会员 = clean_fee(费用_会员_raw)
费用_非会员 = clean_fee(费用_非会员_raw)

# ============================================================
# 三、上课日期转换
# ============================================================
上课日期原文 = str(row["上課日期"]).strip()
if "|" in 上课日期原文:
    parts = 上课日期原文.split("|")
    日期_开始 = re.sub(r"\s*\(開始\)", "", parts[0]).strip()
    日期_结束 = re.sub(r"\s*\(結束\)", "", parts[1]).strip()
    日期 = f"{日期_开始} 至 {日期_结束}"
else:
    日期 = 上课日期原文

# ============================================================
# 四、上课时间转换
# ============================================================
上课时间原文 = str(row["時間"]).strip()
if "|" in 上课时间原文:
    parts = 上课时间原文.split("|")
    start = re.sub(r"\s*\(.*?\)", "", parts[0]).strip()
    end = re.sub(r"\s*\(.*?\)", "", parts[1]).strip()
    上课时间 = f"{start} - {end}"
else:
    上课时间 = 上课时间原文

# ============================================================
# 五、逢星期
# ============================================================
逢星期 = str(row["逢星期"]).strip()

# ============================================================
# 六、堂數字段
# ============================================================
堂數原文 = str(row["堂數"]).strip()


def extract_minutes(text):
    try:
        小時_match = re.search(r"(\d+)\s*小時", text)
        分_match = re.search(r"(\d+)\s*分", text)
        h = int(小時_match.group(1)) if 小時_match else 0
        m = int(分_match.group(1)) if 分_match else 0
        return f"{h*60 + m}分鐘"
    except:
        return ""


def extract_lessons(text):
    try:
        return re.split(r"\(", text)[0].strip()
    except:
        return ""


课程时数 = extract_minutes(堂數原文)
堂数 = extract_lessons(堂數原文)

# ============================================================
# 七、教师姓名清洗
# ============================================================
教師原文 = str(row["導師"]).strip()
教师姓名 = re.split(r"\|\(", 教師原文)[0].strip() if "|(" in 教師原文 else 教師原文

# ============================================================
# 八、会员/非会员/总人数统计
# ============================================================
if "會員類別" in df_receipt.columns:
    原始行数 = len(df_receipt)
    总人数 = 原始行数 - 1 if 原始行数 > 1 else 0
    非会员人数值 = (
        df_receipt["會員類別"].astype(str).str.contains("非會員", regex=False).sum()
    )
    会员人数 = 总人数 - 非会员人数值
    非会员人数 = "N/A" if 非会员人数值 == 0 else 非会员人数值
else:
    总人数 = 0
    会员人数 = 0
    非会员人数 = "N/A"

# ============================================================
# 九、优惠人数统计
# ============================================================
if "扣減" in df_receipt.columns:
    扣減列 = df_receipt["扣減"].astype(str).tolist()
    null个数 = sum(1 for x in 扣減列 if x.strip().lower() in ["", "null", "nan"])
    优惠人数 = len(扣減列) - null个数 - 1 if len(扣減列) > 1 else 0
else:
    优惠人数 = 0

# ============================================================
# 十、处理總額列并校验
# ============================================================
总额列_raw = (
    df_receipt["總額"].astype(str).tolist() if "總額" in df_receipt.columns else []
)
总额数据 = []

for x in 总额列_raw:
    x = x.strip()
    if "總額" in x or x == "":
        continue
    x = x.replace("$", "").replace(",", "")
    try:
        num = float(x)
        if num.is_integer():
            num = int(num)
        总额数据.append(num)
    except:
        continue

if len(总额数据) < 2:
    raise ValueError("❌ 收据中總額数据不足，无法进行校验。")

每项数据 = 总额数据[:-1]
总收入 = 总额数据[-1]

if round(sum(每项数据), 2) != round(总收入, 2):
    raise ValueError(
        f"❌ 收据总额校验失败：明细总和为 {sum(每项数据)}，但总额为 {总收入}"
    )

频率统计 = Counter(每项数据)
表达式_parts = []
for 金额, 次数 in sorted(频率统计.items()):
    金额_str = f"${金额}"
    if 次数 > 1:
        表达式_parts.append(f"{金额_str}×{次数}")
    else:
        表达式_parts.append(f"{金额_str}")
课程总收入计算 = " + ".join(表达式_parts) + f" = ${总收入}"

# ============================================================
# 十一、命令行输入分成信息
# ============================================================
print("\n📌 分成模式选择：")
print("模式 1：课程总收入 × 分成百分比")
print("模式 2：堂数 × 每堂分成费用")
print("模式 3：学生人数 × 每人分成费用")

模式选择 = input("👉 请选择当前课程的分成模式 (输入1/2/3)：").strip()
分成结果 = ""

if 模式选择 == "1":
    百分比值 = float(input("请输入分成百分比(0~100)：").strip())

    # 提取"其他收费"列的最后一项作为"其他收费"金额
    if "其他收費" in df_receipt.columns:
        其他收费列 = df_receipt["其他收費"].dropna().astype(str).tolist()
        try:
            最后值 = 其他收费列[-1].replace("$", "").replace(",", "").strip()
            其他收费值 = float(最后值) if 最后值 else 0
        except:
            其他收费值 = 0
    else:
        其他收费值 = 0

    # 格式化金额（整数不带小数点）
    def format_money(m):
        return f"{int(m)}" if m == int(m) else f"{m:.2f}"

    总收入_fmt = format_money(总收入)
    其他收费_fmt = format_money(其他收费值)

    if 其他收费值 == 0:
        # 无其他收费
        分成金额 = round(总收入 * 百分比值 / 100, 2)
        分成金额_fmt = format_money(分成金额)
        导师费用计算 = f"${总收入_fmt}×{百分比值:.0f}%=${分成金额_fmt}"
        分成结果 = f"分成模式：總收入 × {百分比值:.0f}% = ${分成金额_fmt}"
    else:
        # 有其他收费，先扣除再分成
        分成金额 = round((总收入 - 其他收费值) * 百分比值 / 100, 2)
        分成金额_fmt = format_money(分成金额)
        导师费用计算 = (
            f"(${总收入_fmt}-${其他收费_fmt})×{百分比值:.0f}%=${分成金额_fmt}"
        )
        分成结果 = f"分成模式：(${总收入_fmt}-${其他收费_fmt}) × {百分比值:.0f}% = ${分成金额_fmt}"

    分成 = f"{百分比值:.0f}%"

elif 模式选择 == "2":
    每堂费用 = float(input("请输入每堂分成费用：").strip())
    分成金额 = round(每堂费用 * float(堂数), 2)
    每堂费用_str = f"{int(每堂费用)}" if 每堂费用.is_integer() else f"{每堂费用:.2f}"
    分成金额_str = f"{int(分成金额)}" if 分成金额.is_integer() else f"{分成金额:.2f}"
    分成结果 = f"分成模式：{堂数}堂 × ${每堂费用_str} = ${分成金额_str}"
    导师费用计算 = f"${每堂费用_str}×{堂数}=${分成金额_str}"
    分成 = f"${每堂费用_str}/堂"
elif 模式选择 == "3":
    每人费用 = float(input("请输入每人分成费用：").strip())
    分成金额 = round(每人费用 * 总人数, 2)

    每人费用_str = f"{int(每人费用)}" if 每人费用.is_integer() else f"{每人费用:.2f}"
    分成金额_str = f"{int(分成金额)}" if 分成金额.is_integer() else f"{分成金额:.2f}"

    分成结果 = f"分成模式：{总人数}人 × ${每人费用_str} = ${分成金额_str}"
    导师费用计算 = f"${每人费用_str}×{总人数}=${分成金额_str}"
    分成 = f"${每人费用_str}/人"
else:
    raise ValueError("❌ 分成模式输入错误，请输入 1/2/3。")

中心收入 = 总收入 - 分成金额
中心收入_str = f"{int(中心收入)}" if 中心收入.is_integer() else f"{中心收入:.2f}"
总收入_str = f"{int(总收入)}" if 总收入 == int(总收入) else f"{总收入:.2f}"
分成金额_str = f"{int(分成金额)}" if 分成金额.is_integer() else f"{分成金额:.2f}"
中心收入计算 = f"${总收入_str}-${分成金额_str}=${中心收入_str}"


# ============================================================
# 十二、构造模板上下文
# ============================================================
context = {
    "课程_课程名称": 课程名称_cleaned,
    "课程_课程编号": 课程编号,
    "费用_会员": 费用_会员,
    "费用_非会员": 费用_非会员,
    "日期": 日期,
    "时间": 上课时间,
    "逢星期": 逢星期,
    "课程时数": 课程时数,
    "堂数": 堂数,
    "教师姓名": 教师姓名,
    "总人数": 总人数,
    "会员人数": 会员人数,
    "非会员人数": 非会员人数,
    "优惠人数": 优惠人数,
    "课程总收入计算": 课程总收入计算,
    "分成": 分成结果,
    "分成": 分成,
    "导师费用计算": 导师费用计算,
    # '中心收入计算': f"${总收入}-${分成金额}=${总收入 - 分成金额}",
    "中心收入计算": 中心收入计算,
    "制表日期": datetime.datetime.now().strftime("%d/%m/%Y"),
}

# ============================================================
# 十三、渲染 Word 模板
# ============================================================
try:
    doc = DocxTemplate(template_path)
    doc.render(context)
except Exception as e:
    print("❌ Word 模板渲染失败，请检查字段命名。")
    raise e

# ============================================================
# 十四、输出 Word 文件
# ============================================================
output_filename = f"{课程名称_cleaned}{课程编号}@行政清单.docx"
output_path = os.path.join(output_dir, output_filename)
doc.save(output_path)

print("\n✅ 文件生成成功！")
print(f"📄 输出文件：{output_path}")
print(
    f"📊 统计信息：总人数={总人数} | 会员人数={会员人数} | 非会员人数={非会员人数} | 优惠人数={优惠人数}"
)
print(f"🧾 总额计算表达式：{课程总收入计算}")
print(f"💰 分成结果：{分成结果}")

# ============================================================
# 十五、追加数据到"历史清单汇总.xlsx"
# ============================================================

excel_path = os.path.join(output_dir, "历史清单汇总.xlsx")
sheet_name = "清单数据"

# 字段标题（新增"日期"字段）
excel_fields = [
    "教师姓名",
    "课程_课程名称",
    "课程_课程编号",
    "分成",
    "课程总收入",
    "导师费用",
    "中心收入",
    "日期",
]

# 提取数值并格式化（保留两位小数，无$符号）
总收入_数值 = round(float(总收入), 2)
分成金额_数值 = round(float(分成金额), 2)
中心收入_数值 = round(float(中心收入), 2)

# 构造一行数据（同步添加"日期"字段）
row_data = [
    教师姓名,
    context["课程_课程名称"],
    context["课程_课程编号"],
    context["分成"],
    f"{总收入_数值:.2f}",
    f"{分成金额_数值:.2f}",
    f"{中心收入_数值:.2f}",
    context["日期"],
]

# 判断文件是否存在，处理表格
if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 1 and ws.max_column == len(excel_fields) - 1:
            # 如果旧文件缺少"日期"字段，追加一列标题
            ws.cell(row=1, column=len(excel_fields)).value = "日期"
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(excel_fields)  # 添加表头
else:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(excel_fields)  # 添加表头

# 写入数据行
ws.append(row_data)
wb.save(excel_path)

print(f"📚 已追加数据至历史清单汇总：{excel_path}")

# ============================================================
# 十六、根据历史数据生成该教师的领款单（Word 模板注入）
# ============================================================

# 读取历史清单数据
try:
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
except Exception as e:
    print("❌ 无法读取历史清单汇总，请检查 Excel 文件格式。")
    raise e

# 提取所有教师匹配的数据行（包括本次写入的数据）
rows = list(ws.iter_rows(min_row=2, values_only=True))
matched_rows = [r for r in rows if r[0] == 教师姓名]

if not matched_rows:
    raise ValueError(f"❌ 历史数据中未找到教师 {教师姓名} 的记录，无法生成领款单。")

# 构建"項目名字&編號"字段（多项换行）
项目名字编号列表 = [f"{row[1]}({row[2]})" for row in matched_rows]
项目名字编号_str = "\n".join(项目名字编号列表)

# 计算"項目金額"字段（导师费用总和）
导师费用总和 = sum([float(row[5]) for row in matched_rows])
项目金额_str = f"${导师费用总和:.2f}"


# 金额转换为中文大写（参考VBA逻辑）
def convert_to_chinese_currency(num):
    digits = "零壹貳叁肆伍陸柒捌玖"
    units = ["", "拾", "佰", "仟"]
    big_units = ["", "万", "亿", "兆"]
    decimal_units = ["角", "分"]

    if num < 0:
        return "负" + convert_to_chinese_currency(-num)

    num_str = f"{num:.2f}"
    integer_part, decimal_part = num_str.split(".")
    integer_part = integer_part.lstrip("0") or "0"

    result = ""
    integer_part = integer_part[::-1]
    for i in range(0, len(integer_part), 4):
        group = integer_part[i : i + 4]
        group_str = ""
        zero_flag = False
        for j in range(len(group)):
            n = int(group[j])
            if n == 0:
                if not zero_flag and group_str:
                    group_str = digits[0] + group_str
                zero_flag = True
            else:
                group_str = digits[n] + units[j] + group_str
                zero_flag = False
        group_str = group_str.rstrip(digits[0])
        if group_str:
            result = group_str + big_units[i // 4] + result

    result = result or digits[0]
    result += "元"

    if decimal_part == "00":
        result += "正"
    else:
        jiao = int(decimal_part[0])
        fen = int(decimal_part[1])
        if jiao != 0:
            result += digits[jiao] + decimal_units[0]
        if fen != 0:
            result += digits[fen] + decimal_units[1]

    return result


中文大写金额 = convert_to_chinese_currency(导师费用总和)

# 生成日期相关字段
today = datetime.datetime.today()
day = today.day
month = today.month
year = today.year

if day <= 15:
    领款日期 = f"15/{month}/{year}"
    待输入月份 = month
    期 = "2"
else:
    if month == 12:
        next_month = 1
        next_year = year + 1
    else:
        next_month = month + 1
        next_year = year
    领款日期 = f"1/{next_month}/{next_year}"
    待输入月份 = next_month
    期 = "1"

m1 = str(待输入月份 // 10)
m2 = str(待输入月份 % 10)

# 构建模板上下文
context_lingkuan = {
    "领款日期": 领款日期,
    "m1": m1,
    "m2": m2,
    "期": 期,
    "抬头": 教师姓名,
    "项目名字编号": 项目名字编号_str,
    "项目金额": 项目金额_str,
    "港币圆数大写": 中文大写金额,
}

# 渲染模板文件
# template_lingkuan_path = "领款单-模板.docx"
output_lingkuan_path = os.path.join(output_dir, f"{教师姓名}-领款单.docx")

try:
    doc_lingkuan = DocxTemplate(template_lingkuan_path)
    doc_lingkuan.render(context_lingkuan)
    doc_lingkuan.save(output_lingkuan_path)
except Exception as e:
    print("❌ 领款单模板渲染失败，请检查模板字段是否一致。")
    raise e

print(f"📄 已生成教师领款单：{output_lingkuan_path}")
# ============================================================
# 十七、根据"历史清单汇总.xlsx"生成"支出賬"Excel汇入记录
# ============================================================

# 使用在文件顶部已定义的路径变量
summary_path = excel_path
template_output_path = template_output_spending_actual_path
target_sheet = "支出賬"

if os.path.exists(summary_path) and os.path.exists(template_output_path):
    df_summary = pd.read_excel(summary_path, sheet_name="清单数据")
    grouped = {}

    def parse_date_range(date_str):
        try:
            parts = date_str.split("至")
            start = datetime.datetime.strptime(parts[0].strip(), "%Y-%m-%d")
            end = datetime.datetime.strptime(parts[1].strip(), "%Y-%m-%d")
            return start, end
        except:
            return None, None

    for _, row in df_summary.iterrows():
        name = row["教师姓名"]
        if name not in grouped:
            grouped[name] = {
                "教师姓名": name,
                "课程名称": [],
                "课程编号": [],
                "导师费用": 0,
                "日期范围": [],
            }
        grouped[name]["课程名称"].append(row["课程_课程名称"])
        grouped[name]["课程编号"].append(row["课程_课程编号"])
        grouped[name]["导师费用"] += float(row["导师费用"])
        grouped[name]["日期范围"].append(parse_date_range(row["日期"]))

    today = datetime.datetime.today()
    if today.day <= 15:
        f_column = f"{today.year}-{today.month:02d}-15"
    else:
        if today.month == 12:
            f_column = f"{today.year + 1}-01-01"
        else:
            f_column = f"{today.year}-{today.month + 1:02d}-01"
    i_column = today.strftime("%Y-%m-%d")

    data_rows = []
    counter = 1
    for name, info in grouped.items():
        date_ranges = [r for r in info["日期范围"] if r[0] and r[1]]
        if date_ranges:
            start = min(r[0] for r in date_ranges)
            end = max(r[1] for r in date_ranges)
            date_range_str = (
                f"{start.strftime('%Y-%m-%d')} 至 {end.strftime('%Y-%m-%d')}"
            )
        else:
            date_range_str = ""

        data_rows.append(
            [
                counter,
                "T005",
                "山景-SK",
                "---",
                "---",
                f_column,
                "C029",
                "導師費",
                i_column,
                "/".join(info["课程名称"]),
                round(info["导师费用"], 2),
                "否",
                "否",
                name,
                "/".join(info["课程编号"]),
                date_range_str,
            ]
        )
        counter += 1

    wb_out = load_workbook(template_output_path)
    ws_out = wb_out[target_sheet]

    start_row = 9
    for i, row in enumerate(data_rows):
        for j, value in enumerate(row):
            col_letter = get_column_letter(j + 1)
            ws_out[f"{col_letter}{start_row + i}"] = value

    wb_out.save(template_output_path)
    print(f"📥 已成功写入支出賬表格：{template_output_path}")
    input("\n按 Enter 键退出...")
else:
    print("⚠️ 缺少历史清单或模板文件，已跳过支出賬写入。")
