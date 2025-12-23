# -*- coding: utf-8 -*-
import os
import sys
import shutil
import datetime
import re
import fitz  # PyMuPDF
from PyPDF2 import PdfReader
from docxtpl import DocxTemplate
from openpyxl import load_workbook

# ============================================================
# ⚙️ 路径配置 (相对路径)
# ============================================================
# 当前脚本所在目录 (1B_杂费领款单)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 项目根目录 (测试整理目录)
ROOT_DIR = os.path.dirname(BASE_DIR)

# 配置文件目录 (用于引用 config_paths, 也可以直接在此定义)
CONFIG_DIR = os.path.join(ROOT_DIR, "0_模板文件及初始化")
sys.path.append(CONFIG_DIR)

# 输入文件夹
INPUT_DIR_PRINT = os.path.join(BASE_DIR, "此处放入打印费文件")
INPUT_DIR_NET = os.path.join(BASE_DIR, "此处放入上网费文件")
INPUT_DIR_FB = os.path.join(BASE_DIR, "此处放入FaceBook宣传费文件")

# 输出文件夹
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

# 已处理归档文件夹
ARCHIVE_DIR_ROOT = os.path.join(BASE_DIR, "已处理文件")
today_str = datetime.datetime.now().strftime("%Y%m%d")
ARCHIVE_DIR_TODAY = os.path.join(ARCHIVE_DIR_ROOT, today_str)

# 模板文件路径
TEMPLATE_HP = os.path.join(CONFIG_DIR, "HP Inc Hong Kong Limited.docx")
TEMPLATE_NET = os.path.join(CONFIG_DIR, "Information Technology Resource Centre.docx")
TEMPLATE_FB = os.path.join(CONFIG_DIR, "Knight Creative Limited模板.docx")

# Excel 记账文件路径
EXCEL_PATH = os.path.join(ROOT_DIR, "2_Excel滙入記錄模板-支出賬文件", "屯門婦聯 - 會計及財務記賬系統-Excel滙入記錄模板-支出賬.xlsx")

# 确保输出目录存在
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ============================================================
# 🛠️ 通用工具函数
# ============================================================

def convert_to_chinese_currency(num):
    """数字转中文大写"""
    digits = "零壹貳叁肆伍陸柒捌玖"
    units = ["", "拾", "佰", "仟"]
    big_units = ["", "萬", "億", "兆"]
    decimal_units = ["角", "分"]

    try:
        num = float(num)
    except:
        return "零元正"

    if num < 0:
        return "负" + convert_to_chinese_currency(-num)

    num_str = f"{num:.2f}"
    integer_part, decimal_part = num_str.split('.')
    integer_part = integer_part.lstrip('0') or '0'

    result = ""
    integer_part = integer_part[::-1]
    for i in range(0, len(integer_part), 4):
        group = integer_part[i:i+4]
        group_str = ""
        zero_flag = False
        for j in range(len(group)):
            n = int(group[j])
            if n == 0:
                if not zero_flag and group_str:
                    group_str = digits[0] + group_str
                zero_flag = True
            else:
                group_str = digits[n] + units[j] + group_str
                zero_flag = False
        group_str = group_str.rstrip(digits[0])
        if group_str:
            result = group_str + big_units[i // 4] + result

    result = result or digits[0]
    result += "元"

    if decimal_part == "00":
        result += "正"
    else:
        jiao = int(decimal_part[0])
        fen = int(decimal_part[1])
        if jiao != 0:
            result += digits[jiao] + decimal_units[0]
        if fen != 0:
            result += digits[fen] + decimal_units[1]

    return result

def get_date_logic():
    """
    生成领款单需要的日期字段 (m_date, m1, m2, 期)
    以及 Excel 需要的生效日期 (effective_date)
    """
    today = datetime.datetime.today()
    day = today.day
    month = today.month
    year = today.year
    
    # 逻辑：当前日期 <= 15 -> 本月15号； > 15 -> 下月1号
    if day <= 15:
        # Word用
        w_date = f"15/{month}/{year}"
        w_month_used = month
        w_period = "2"
        # Excel用 (本月15日)
        e_date = datetime.datetime(year, month, 15)
    else:
        if month == 12:
            w_month_used = 1
            year_next = year + 1
        else:
            w_month_used = month + 1
            year_next = year
            
        # Word用
        w_date = f"1/{w_month_used}/{year_next}"
        w_period = "1"
        # Excel用 (下月1日)
        e_date = datetime.datetime(year_next, w_month_used, 1)

    return {
        "word_date": w_date,
        "m1": str(w_month_used // 10),
        "m2": str(w_month_used % 10),
        "period": w_period,
        "excel_effective_date": e_date,
        "run_date": today # 脚本运行日期
    }

def move_file_to_archive(file_path):
    """处理完成后将文件移动到归档目录"""
    if not os.path.exists(ARCHIVE_DIR_TODAY):
        os.makedirs(ARCHIVE_DIR_TODAY)
    filename = os.path.basename(file_path)
    shutil.move(file_path, os.path.join(ARCHIVE_DIR_TODAY, filename))
    print(f"📦 文件已归档至: {os.path.join(ARCHIVE_DIR_TODAY, filename)}")

# ============================================================
# 📄 PDF 解析逻辑 (打印费/网费 - PyPDF2)
# ============================================================
def process_print_file(pdf_path):
    """处理打印费 PDF"""
    reader = PdfReader(pdf_path)
    text = "".join(page.extract_text() for page in reader.pages)

    # 提取 Invoice Number
    inv_match = re.search(r"Invoice Number\s+(\d+)", text)
    invoice_no = inv_match.group(1) if inv_match else "Unknown"

    # 提取 Total Amount
    amt_match = re.search(r"Total Amount\s+([0-9]+\.[0-9]{2})", text)
    amount = float(amt_match.group(1)) if amt_match else 0.0

    # 提取 Clicks Charge Period (用于项目名称)
    # 格式: 29 May 2025 - 28 Jun 2025 -> 取结束日期的月份
    date_match = re.search(r"Clicks Charge Period\s+([0-9]{2} \w{3} [0-9]{4})\s*-\s*([0-9]{2} \w{3} [0-9]{4})", text)
    
    month_str, year_str = "", ""
    if date_match:
        end_date_str = date_match.group(2) # e.g. "28 Jun 2025"
        dt_obj = datetime.datetime.strptime(end_date_str, "%d %b %Y")
        month_str = str(dt_obj.month)
        year_str = str(dt_obj.year)
    else:
        # Fallback: 使用当前月份
        now = datetime.datetime.now()
        month_str = str(now.month)
        year_str = str(now.year)

    project_name = f"影印費{month_str}/{year_str}(InvoiceNo.{invoice_no})"
    excel_desc = f"影印費({month_str}/{year_str})"

    # 生成数据
    date_info = get_date_logic()
    
    context = {
        "领款日期": date_info["word_date"],
        "m1": date_info["m1"],
        "m2": date_info["m2"],
        "期": date_info["period"],
        "项目名字编号": project_name,
        "项目金额": f"${amount:.2f}",
        "港币圆数大写": convert_to_chinese_currency(amount)
    }

    # Excel 数据包
    excel_data = {
        "type": "PRINT",
        "effective_date": date_info["excel_effective_date"],
        "run_date": date_info["run_date"],
        "desc": excel_desc,
        "amount": amount,
        "invoice_no": invoice_no
    }

    return context, excel_data, TEMPLATE_HP, f"{year_str}年{month_str}月打印费领款单.docx"

def process_net_file(pdf_path):
    """处理上网费 PDF"""
    reader = PdfReader(pdf_path)
    text = "".join(page.extract_text() for page in reader.pages)

    # 提取 Invoice No
    inv_match = re.search(r"INVOICE NO\.\s*:\s*(\d+)", text)
    invoice_no = inv_match.group(1) if inv_match else "Unknown"

    # 提取 Invoice Date
    date_match = re.search(r"INVOICE DATE\s*:\s*(\d{2})/(\d{2})/(\d{4})", text)
    if date_match:
        month = int(date_match.group(2))
        year = int(date_match.group(3))
    else:
        now = datetime.datetime.now()
        month, year = now.month, now.year

    project_name = f"山景中心上網費({month}/{year})(NO.{invoice_no})"
    excel_desc = f"山景中心上網費({month}/{year})"

    date_info = get_date_logic()
    
    context = {
        "领款日期": date_info["word_date"],
        "m1": date_info["m1"],
        "m2": date_info["m2"],
        "期": date_info["period"],
        "项目名字编号": project_name
    }

    # Excel 数据包 - 注意金额固定478
    excel_data = {
        "type": "NET",
        "effective_date": date_info["excel_effective_date"],
        "run_date": date_info["run_date"],
        "desc": excel_desc,
        "amount": 478.00, # 固定金额
        "invoice_no": invoice_no
    }

    return context, excel_data, TEMPLATE_NET, f"{year}年{month}月网费领款单.docx"

# ============================================================
# 📘 PDF 解析逻辑 (Facebook - PyMuPDF/Fitz)
# ============================================================
def process_fb_file(pdf_path):
    """处理 Facebook PDF"""
    doc = fitz.open(pdf_path)
    target_page = None
    invoice_number = None
    hkd_amount = None

    for page in doc:
        text = page.get_text()
        if "山景服務處" in text:
            target_page = page
            break
    
    if not target_page:
        raise ValueError("❌ 未找到 '山景服務處' 页面")

    # 提取文本块
    lines = []
    blocks = target_page.get_text("dict")["blocks"]
    for b in blocks:
        for line in b.get("lines", []):
            line_text = " ".join(span["text"].strip() for span in line["spans"])
            lines.append(line_text.strip())

    # 提取金额
    balance_indices = [i for i, l in enumerate(lines) if l == "Balance Due"]
    for idx in balance_indices:
        if idx + 1 < len(lines):
            next_line = lines[idx + 1].strip()
            if next_line.startswith("HKD"):
                hkd_amount = next_line
                break
    
    if not hkd_amount:
        # 尝试备用提取逻辑，有时候金额在同一行
        raise ValueError("❌ 未找到金额 (Balance Due)")
    
    amount_clean = hkd_amount.replace("HKD", "").replace(",", "").strip()
    amount_float = float(amount_clean)

    # 提取发票号 (Project ID)
    for line in lines:
        if line.startswith("# INV-"):
            invoice_number = line
            break
    
    if not invoice_number:
        invoice_number = "Unknown"
    
    project_id = invoice_number.replace("# ", "").strip()
    
    # 日期逻辑 (通常取当前日期作为描述中的日期)
    now = datetime.datetime.now()
    excel_desc = f"網上宣傳費({now.month}/{now.year})"

    date_info = get_date_logic()

    context = {
        "项目金额": f"${amount_float:,.2f}",
        "项目编号": project_id,
        "港币圆数大写": convert_to_chinese_currency(amount_float),
        "领款日期": date_info["word_date"],
        "m1": date_info["m1"],
        "m2": date_info["m2"],
        "期": date_info["period"],
    }

    excel_data = {
        "type": "FB",
        "effective_date": date_info["excel_effective_date"],
        "run_date": date_info["run_date"],
        "desc": excel_desc,
        "amount": amount_float,
        "invoice_no": project_id
    }

    return context, excel_data, TEMPLATE_FB, "FaceBook宣传费领款单.docx"

# ============================================================
# 📊 Excel 写入逻辑
# ============================================================
def append_to_excel(data_list):
    if not data_list:
        return

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel 文件不存在: {EXCEL_PATH}")
        return

    print(f"🔄 正在写入 Excel ({len(data_list)} 条记录)...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb["支出賬"]

    # 寻找第9行开始的第一个空行
    start_row = 9
    current_row = start_row
    
    # 简单的寻找最大序号逻辑
    max_seq = 0
    
    # 遍历寻找空行同时记录最大序号
    while ws[f"A{current_row}"].value is not None:
        val = ws[f"A{current_row}"].value
        if isinstance(val, int):
            if val > max_seq:
                max_seq = val
        current_row += 1
    
    for data in data_list:
        max_seq += 1
        r = current_row
        
        # 提取通用数据
        eff_date = data["effective_date"] # datetime object
        run_date_str = data["run_date"].strftime("%Y-%m-%d")
        
        # 写入通用列
        ws[f"A{r}"] = max_seq          # 序号
        ws[f"B{r}"] = "T005"           # 固定
        ws[f"C{r}"] = "山景-SK"        # 固定
        ws[f"D{r}"] = "---"
        ws[f"E{r}"] = "---"
        ws[f"F{r}"] = eff_date         # 生效日期 (Excel会处理datetime)
        ws[f"F{r}"].number_format = 'yyyy-mm-dd'
        
        ws[f"I{r}"] = run_date_str     # 录入日期
        ws[f"L{r}"] = "否"
        ws[f"M{r}"] = "否"

        # 写入特定列
        d_type = data["type"]
        
        if d_type == "PRINT":
            ws[f"G{r}"] = "C021"
            ws[f"H{r}"] = "印刷"
            ws[f"J{r}"] = data["desc"]
            ws[f"K{r}"] = data["amount"]
            ws[f"N{r}"] = "HP Inc Hong Kong Limited"
            ws[f"O{r}"] = data["invoice_no"]
            
        elif d_type == "NET":
            ws[f"G{r}"] = "C025"
            ws[f"H{r}"] = "電話及互聯網費"
            ws[f"J{r}"] = data["desc"]
            ws[f"K{r}"] = data["amount"] # 478
            ws[f"N{r}"] = "Information Technology Resource Centre"
            ws[f"O{r}"] = data["invoice_no"]
            
        elif d_type == "FB":
            ws[f"G{r}"] = "C013"
            ws[f"H{r}"] = "廣告及推廣"
            ws[f"J{r}"] = data["desc"]
            ws[f"K{r}"] = data["amount"]
            ws[f"N{r}"] = "Knight Creative Limited"
            ws[f"O{r}"] = data["invoice_no"]

        current_row += 1

    wb.save(EXCEL_PATH)
    print("✅ Excel 写入完成。")

# ============================================================
# 🚀 主程序
# ============================================================
def main():
    excel_queue = []

    # 1. 扫描 打印费
    if os.path.exists(INPUT_DIR_PRINT):
        for f in os.listdir(INPUT_DIR_PRINT):
            if f.lower().endswith(".pdf"):
                f_path = os.path.join(INPUT_DIR_PRINT, f)
                try:
                    print(f"🖨️ 正在处理打印费: {f}")
                    ctx, xls_data, tpl_path, out_name = process_print_file(f_path)
                    
                    doc = DocxTemplate(tpl_path)
                    doc.render(ctx)
                    doc.save(os.path.join(OUTPUT_DIR, out_name))
                    
                    excel_queue.append(xls_data)
                    move_file_to_archive(f_path)
                except Exception as e:
                    print(f"❌ 处理打印费 {f} 失败: {e}")

    # 2. 扫描 上网费
    if os.path.exists(INPUT_DIR_NET):
        for f in os.listdir(INPUT_DIR_NET):
            if f.lower().endswith(".pdf"):
                f_path = os.path.join(INPUT_DIR_NET, f)
                try:
                    print(f"🌐 正在处理上网费: {f}")
                    ctx, xls_data, tpl_path, out_name = process_net_file(f_path)
                    
                    doc = DocxTemplate(tpl_path)
                    doc.render(ctx)
                    doc.save(os.path.join(OUTPUT_DIR, out_name))
                    
                    excel_queue.append(xls_data)
                    move_file_to_archive(f_path)
                except Exception as e:
                    print(f"❌ 处理上网费 {f} 失败: {e}")

    # 3. 扫描 Facebook
    if os.path.exists(INPUT_DIR_FB):
        for f in os.listdir(INPUT_DIR_FB):
            if f.lower().endswith(".pdf"):
                f_path = os.path.join(INPUT_DIR_FB, f)
                try:
                    print(f"📘 正在处理 FB 宣传费: {f}")
                    ctx, xls_data, tpl_path, out_name = process_fb_file(f_path)
                    
                    doc = DocxTemplate(tpl_path)
                    doc.render(ctx)
                    doc.save(os.path.join(OUTPUT_DIR, out_name))
                    
                    excel_queue.append(xls_data)
                    move_file_to_archive(f_path)
                except Exception as e:
                    print(f"❌ 处理 FB {f} 失败: {e}")

    # 4. 写入 Excel
    if excel_queue:
        append_to_excel(excel_queue)
    else:
        print("ℹ️ 未发现新文件，无需更新 Excel。")

if __name__ == "__main__":
    main()