import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font

# 读取 JSON 文件
with open("get_attendance.json", "r", encoding="utf-8") as f:
    data = json.load(f)

# 提取 membership 数据
members = data.get("data", {}).get("membership", [])

# 初始化列表
residence_data = []
mobile_data = []

# 遍历会员信息
for member in members:
    name_zh = member.get("name_zh", "").strip()
    contact_residence = member.get("contact_residence", "").strip()
    contact_mobile = member.get("contact_mobile", "").strip()

    if contact_residence:
        residence_data.append([name_zh, contact_residence])
    if contact_mobile:
        mobile_data.append([name_zh, contact_mobile])

# 写入初步 Excel 文件
excel_filename = "contact.xlsx"
with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
    pd.DataFrame(mobile_data, columns=["姓名", "手机"]).to_excel(writer, sheet_name="手机", index=False)
    pd.DataFrame(residence_data, columns=["姓名", "住宅电话"]).to_excel(writer, sheet_name="住宅电话", index=False)

# 加载 Excel 并添加超链接公式
wb = load_workbook(excel_filename)

def add_hyperlinks(sheet_name, phone_column_name):
    sheet = wb[sheet_name]
    col_B = 2  # 电话列（B列）
    col_C = 3  # 要插入超链接的C列

    # 写入标题
    sheet.cell(row=1, column=col_C, value="WhatsApp链接")

    # 写入每一行超链接函数
    for row in range(2, sheet.max_row + 1):
        phone = sheet.cell(row=row, column=col_B).value
        if phone:
            formula = f'=HYPERLINK("https://wa.me/852"&B{row}, "👉 點此發送訊息")'
            sheet.cell(row=row, column=col_C, value=formula)

# 分别处理两个表
add_hyperlinks("手机", "手机")
add_hyperlinks("住宅电话", "住宅电话")

# 保存修改后的文件
wb.save(excel_filename)
print("✅ 已成功写入 contact.xlsx，并添加 WhatsApp 链接")
