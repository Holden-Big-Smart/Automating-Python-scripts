import json
import pandas as pd
from openpyxl import load_workbook

# 读取 JSON 文件
with open("get_attendance.json", "r", encoding="utf-8") as f:
    data = json.load(f)

# 提取 membership 数据
members = data.get("data", {}).get("membership", [])

# 初始化统一联系人列表
contact_list = []

# 遍历会员信息，提取住宅电话和手机
for member in members:
    name = member.get("name_zh", "").strip()
    residence = member.get("contact_residence", "").strip()
    mobile = member.get("contact_mobile", "").strip()

    if residence:
        contact_list.append([name, residence])
    if mobile:
        contact_list.append([name, mobile])

# 写入 Excel（初步 DataFrame 保存）
excel_filename = "contact.xlsx"
df = pd.DataFrame(contact_list, columns=["姓名", "电话"])
df.to_excel(excel_filename, index=False)

# 打开 Excel 文件并添加 C列超链接
wb = load_workbook(excel_filename)
ws = wb.active  # 默认只有一个工作表

# 写入 C1 标题
ws.cell(row=1, column=3, value="WhatsApp链接")

# 从第二行起，为每一行添加超链接公式
for row in range(2, ws.max_row + 1):
    formula = f'=HYPERLINK("https://wa.me/852"&B{row}, "👉 點此發送訊息")'
    ws.cell(row=row, column=3, value=formula)

# 保存修改后的文件
wb.save(excel_filename)
print("✅ 已生成 contact.xlsx，并添加 WhatsApp 超链接")
