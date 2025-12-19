import pyautogui
import openpyxl
import os
import time
import random
import keyboard
import cv2
import numpy as np
import threading

from utils.trymatch import 截图匹配
from utils.trymatchloop import 循环截图匹配


# -----------------------------
# 全局紧急停止线程：数字小键盘 0 立即强退
# -----------------------------
def emergency_kill_listener():
    while True:
        if keyboard.is_pressed("num 0"):
            print("\n[紧急中止] 检测到 num 0，强制退出脚本")
            os._exit(0)


listener_thread = threading.Thread(target=emergency_kill_listener, daemon=True)
listener_thread.start()

# -----------------------------
# 路径设置（统一用相对路径）
# -----------------------------
base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

contact_path = os.path.join(
    base_dir, "Whatsapp發送腳本", "send_info", "input", "contact.xlsx"
)
uncontact_path = os.path.join(
    base_dir, "Whatsapp發送腳本", "send_info", "output", "uncontact.xlsx"
)
template_dialogue = os.path.join(base_dir, "project", "screenshot", "new_dialogue.png")
template_unfind = os.path.join(base_dir, "project", "screenshot", "unfind.png")


# -----------------------------
# 功能函数
# -----------------------------
# 打开excel
def load_workbook_safe(path):
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        wb.save(path)
        return wb


# 粘贴电话号码
def paste_text(text):
    pyautogui.hotkey("ctrl", "a")
    pyautogui.typewrite(text)
    print(f"[✓] 已粘贴号码：{text}")


# 处理失败联系人
def append_to_uncontact(name, phone):
    wb = load_workbook_safe(uncontact_path)
    ws = wb.active
    row = ws.max_row + 1 if ws["A1"].value else 1
    ws.cell(row=row, column=1).value = name
    ws.cell(row=row, column=2).value = phone
    wb.save(uncontact_path)
    print(f"[×] 联系失败，已保存至 uncontact.xlsx：{name} - {phone}")


# 删除已处理行
def delete_contact_row(wb, ws, row):
    ws.delete_rows(row)
    wb.save(contact_path)
    print(f"[✓] 已从 contact.xlsx 删除已处理行")


# -----------------------------
# 主执行逻辑
# -----------------------------
def main():
    print("[启动] Whatsapp 自动发送脚本")
    print(f"[路径] 项目根目录：{base_dir}")
    print(f"[路径] 联系人文件：{contact_path}")
    print(f"[路径] 模板截图：{template_dialogue}")
    print("[提示] 按下 ESC 键可随时中止程序 | num 0 可全局强制中止")
    print("---------------------------------------------------")
    print("請最小化命令窗口")
    time.sleep(1)
    print("請最小化命令窗口")
    time.sleep(2)
    print("請最小化命令窗口")
    time.sleep(3)
    print("請最小化命令窗口")

    while True:
        if keyboard.is_pressed("esc"):
            print("[用户中止] 检测到 ESC 键，程序已终止")
            break

        if not os.path.exists(contact_path):
            print(f"[错误] 联系人文件不存在：{contact_path}")
            break

        wb = openpyxl.load_workbook(contact_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            name = row[0].value
            phone = row[1].value

            if phone is None:
                print("[完成] 无更多联系人待处理")
                return

            contact_num = str(phone)
            print(f"\n[处理中] 联系人：{name} - {contact_num}")

            time.sleep(1)

            # 匹配新增联系人
            结果 = 循环截图匹配(0, 0, 1365, 767, "assets/new_dialogue.png", 0.9, 1)
            x, y = 结果

            # 点击新增联系人
            pyautogui.click(x, y)

            while True:
                结果 = 截图匹配(0, 0, 1365, 767, "assets/new_dialogue_tittle.png", 0.9)
                if 结果:
                    break
                else:
                    pyautogui.click(x, y)
                    time.sleep(1)
                
            # 粘贴电话
            paste_text(contact_num)
            time.sleep(1)
            # 循环检测是否仍然存在'新增群组'
            while True:
                result = 截图匹配(0, 0, 1365, 767, "assets/loadingfind.png", 0.9)

                if result:
                    print("[✓] 匹配成功，坐标：", result)
                    time.sleep(1)  # 每次间隔1秒，防止刷屏太快
                else:
                    print("[×] 未找到目标图像，退出循环")
                    break

            # 检测是否为'找不到联系人'
            结果 = 截图匹配(0, 0, 1365, 767, "assets/unfind.png", 0.9)
            # 如果出现'找不到联系人'，则加入联系失败列表
            if 结果:
                print("[×] 未找到联系人")
                pyautogui.click(1055, 30)
                time.sleep(1)
                append_to_uncontact(name, contact_num)
                delete_contact_row(wb, ws, row[0].row)
                # 循环检测是否仍然存在'新增对话'
                while True:
                    result = 截图匹配(
                        0, 0, 1365, 767, "assets/new_dialogue_tittle.png", 0.9
                    )

                    if result:
                        print("[✓] 匹配成功，坐标：", result)
                        time.sleep(1)  # 每次间隔1秒，防止刷屏太快
                        pyautogui.click(1055, 30)
                    else:
                        print("[×] 未找到'新增对话'，退出循环")
                        break
            # 否则执行粘贴任务：
            else:
                
                # ↓↓↓ 新的复制粘贴逻辑开始 ↓↓↓
                time.sleep(1)
                # 点击第一个联系人
                pyautogui.click(1000, 260)  # 点击输入区域
                time.sleep(1)
                # 循环检测是否仍然存在'新增对话'
                while True:
                    result = 截图匹配(
                        0, 0, 1365, 767, "assets/new_dialogue_tittle.png", 0.9
                    )
                    if result:
                        print("[✓] 匹配成功，坐标：", result)
                        time.sleep(1)  # 每次间隔1秒，防止刷屏太快
                        pyautogui.click(1000, 260)
                    else:
                        print("[×] 未找到'新增对话'，退出循环")
                        break
                    
                # 激活待复制文件区
                pyautogui.click(635, 400)  # 点击图片预处理区域
                time.sleep(1)
                # 全选待复制文件区内的文件
                pyautogui.hotkey("ctrl", "a")  # 全选
                time.sleep(1)
                # 复制待复制文件区内的文件
                pyautogui.hotkey("ctrl", "c")  # 复制
                time.sleep(1)
                # 查找并检测'输入讯息是否被遮挡'
                结果 = 循环截图匹配(0, 0, 1365, 767, "assets/inputinfo.png", 0.9, 1)
                x, y = 结果
                # 点击输入框
                pyautogui.click(x, y)
                # 粘贴图片
                pyautogui.hotkey("ctrl", "v")  # 粘贴
                time.sleep(1)
                结果 = 循环截图匹配(0, 0, 1365, 767, "assets/white_send.png", 0.9, 1)
                # 点击发送
                pyautogui.press("enter")
                print("[✓] 图片发送成功")
                delete_contact_row(wb, ws, row[0].row)
                time.sleep(2)
                # ↑↑↑ 新的复制粘贴逻辑结束 ↑↑↑
                time.sleep(random.uniform(1, 3))
            break  # 每次运行处理一个联系人


# -----------------------------
# 程序入口
# -----------------------------
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"[错误] 脚本发生异常：{e}")
