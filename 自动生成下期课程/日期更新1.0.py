import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

# ========== 配置 ==========
BASE_DIR = os.getcwd()
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
HOLIDAY_PATH = os.path.join(BASE_DIR, "节假日安排.xlsx")

# 加载节假日列表
holidays = pd.read_excel(HOLIDAY_PATH, dtype=str)
holiday_dates = set(pd.to_datetime(holidays.iloc[:, 0], errors="coerce").dropna().dt.date)

# 列宽配置
col_widths = {
    "A": 18, "B": 10, "C": 10, "D": 10,
    "E": 10, "F": 10, "G": 10, "H": 20
}

# 日期顺延判断函数
def skip_holidays(start_date):
    while start_date in holiday_dates:
        start_date += timedelta(days=1)
    return start_date

# 处理单个 CSV 文件
def process_csv(csv_path):
    filename = os.path.basename(csv_path).replace(".csv", "")
    xlsx_path = os.path.join(OUTPUT_DIR, filename + ".xlsx")

    # 读取 CSV 文件并写入 Sheet1
    df = pd.read_csv(csv_path, dtype=str).fillna("")
    df.to_excel(xlsx_path, index=False, sheet_name="原始数据")

    # 载入工作簿，准备添加“日期更新”工作表
    wb = load_workbook(xlsx_path)
    ws = wb.create_sheet("日期更新")

    # 写入标题
    headers = ["名稱", "下期开课时间", "下期结课时间", "本期上課日期", "堂數", "導師", "編號", "备注"]
    ws.append(headers)

    for idx, row in df.iterrows():
        name = row.get("名稱", "")
        start_str = row.get("上課日期", "")
        lessons = int(str(row.get("堂數", "0")).split()[0]) if str(row.get("堂數", "0")).isdigit() else 0
        teacher = row.get("導師", "")
        code = row.get("編號", "")

        # 解析原始日期
        start_date = None
        try:
            start_date = pd.to_datetime(start_str[:10], errors="coerce")
        except:
            pass

        end_date = None
        next_start = None
        next_end = None

        remark = ""

        if start_date and lessons > 0:
            end_date = start_date + timedelta(weeks=lessons - 1)
            next_start = skip_holidays(end_date + timedelta(days=7))
            next_end = skip_holidays(next_start + timedelta(weeks=lessons - 1))
        else:
            remark = "原始数据异常"

        # 写入行数据
        ws.append([
            name,
            next_start.strftime("%Y-%m-%d") if next_start else "",
            next_end.strftime("%Y-%m-%d") if next_end else "",
            start_date.strftime("%Y-%m-%d") if start_date else "",
            lessons,
            teacher,
            code,
            remark
        ])

    # 设置列宽
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 居中样式
    for row in ws.iter_rows(min_row=2, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(xlsx_path)
    print(f"✅ 已处理：{os.path.basename(csv_path)} → {os.path.basename(xlsx_path)}")

# ========== 执行入口 ==========
if __name__ == "__main__":
    print("📅 正在处理课程日期更新...")

    for fname in os.listdir(OUTPUT_DIR):
        if fname.endswith(".csv") and "待更新课程" in fname:
            process_csv(os.path.join(OUTPUT_DIR, fname))

    print("🎉 所有文件已处理完毕。")
