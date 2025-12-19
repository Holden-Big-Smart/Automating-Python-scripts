import os
import pandas as pd
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ========= 配置路径 =========
BASE_DIR = os.getcwd()
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

# ========= 固定节假日（不再读取「节假日安排.xlsx」） =========
holiday_dates = {
    datetime(2025, 12, 25).date(),
    datetime(2025, 12, 26).date(),
    datetime(2026, 1, 1).date(),
    datetime(2026, 2, 17).date(),
    datetime(2026, 2, 18).date(),
    datetime(2026, 2, 19).date(),
    datetime(2026, 4, 3).date(),
    datetime(2026, 4, 4).date(),
    datetime(2026, 4, 6).date(),
    datetime(2026, 4, 7).date(),
    datetime(2026, 5, 1).date(),
    datetime(2026, 5, 25).date(),
    datetime(2026, 6, 19).date(),
    datetime(2026, 7, 1).date(),
    datetime(2026, 9, 26).date(),
    datetime(2026, 10, 1).date(),
    datetime(2026, 10, 19).date(),
    datetime(2026, 12, 25).date(),
    datetime(2026, 12, 26).date(),
}

# ========= 工具函数 =========


def extract_lessons(raw):
    """从「堂數」字段中提取堂数，如：'6 (堂數)|1 (停課)|...' → 6"""
    match = re.search(r"(\d+)\s*\(堂數\)", str(raw))
    return int(match.group(1)) if match else 0


def adjust_dates_for_holidays(dates, holiday_dates):
    """
    通用节假日处理：
    - 传入一串按 7 天间隔排好的 datetime 列表 dates（同一堂课的排课表）
    - 若某天是节假日：
        * 记入 conflict_dates
        * 从原列表最后一天 +7 开始往后找非节假日日期，作为补课日期加入 valid_dates
    - 最终 valid_dates 长度与原始 dates 一样，只是节假日被替换成后面顺延的日期
    """
    valid_dates = []
    conflict_dates = []

    if not dates:
        return valid_dates, conflict_dates

    # 从原始列表最后一天 +7 开始准备补课日期
    extra_date = dates[-1] + timedelta(days=7)

    for d in dates:
        if d.date() in holiday_dates:
            # 记录冲突
            conflict_dates.append(d)

            # 找到一个不在节假日的补课日期（同一星期几往后顺延）
            while extra_date.date() in holiday_dates:
                extra_date += timedelta(days=7)

            valid_dates.append(extra_date)
            # 下一次补课再往后推一周
            extra_date += timedelta(days=7)
        else:
            valid_dates.append(d)

    return valid_dates, conflict_dates


# 列宽设置
col_widths = {
    "A": 60,
    "B": 15,
    "C": 15,
    "D": 40,
    "E": 8,
    "F": 10,
    "G": 12,
    "H": 25,
    "I": 255,
}


def process_csv(csv_path):
    filename = os.path.basename(csv_path).replace(".csv", "")
    xlsx_path = os.path.join(OUTPUT_DIR, filename + ".xlsx")

    # 读入原始 CSV 并保存为原始数据工作表
    df = pd.read_csv(csv_path, dtype=str).fillna("")
    df.to_excel(xlsx_path, index=False, sheet_name="原始数据")

    wb = load_workbook(xlsx_path)
    ws = wb.create_sheet("日期更新")

    # 新表表头：
    # 上课日期 = 下期所有上课日期（X 拼接）
    headers = [
        "名稱",
        "下期开课时间",
        "下期结课时间",
        "本期上課日期",
        "堂數",
        "導師",
        "編號",
        "备注",
        "上课日期",
    ]
    ws.append(headers)

    for _, row in df.iterrows():
        name = row.get("名稱", "")
        start_str = str(
            row.get("上課日期", "")
        ).strip()  # 如：'2025-10-16 (開始)|2025-11-27 (結束)'
        teacher = row.get("導師", "")
        code = row.get("編號", "")
        raw_lesson = row.get("堂數", "")
        lessons = extract_lessons(raw_lesson)

        original_date_display = start_str  # 保留原始「上課日期」字段显示
        remark = ""
        next_start_str = ""
        next_end_str = ""
        next_term_dates_str = ""

        # 解析「本期开课日」 = 上課日期前 10 个字符
        try:
            current_start_date = pd.to_datetime(start_str[:10], errors="coerce")
        except Exception:
            current_start_date = None

        # 异常情况：无法解析日期或堂数无效
        if pd.isna(current_start_date) or lessons <= 0:
            remark = "未安排課節"
        else:
            # ========== 第一步：生成「本期」的实际上课日期（支持节假日顺延） ==========
            current_naive_dates = [
                current_start_date + timedelta(days=7 * i) for i in range(lessons)
            ]
            current_valid_dates, _ = adjust_dates_for_holidays(
                current_naive_dates, holiday_dates
            )

            # 本期「真实最后一节课」日期
            last_current_date = current_valid_dates[-1]

            # ========== 第二步：计算「下期开课时间」（你的 B 方案） ==========
            # 必须与最后一节同星期；先 +7 天，再按 7 天跳过节假日
            next_start_date = last_current_date + timedelta(days=7)
            while next_start_date.date() in holiday_dates:
                next_start_date += timedelta(days=7)

            # ========== 第三步：生成「下期所有上课日期」，并应用节假日顺延规则 ==========
            next_naive_dates = [
                next_start_date + timedelta(days=7 * i) for i in range(lessons)
            ]
            next_valid_dates, conflict_dates = adjust_dates_for_holidays(
                next_naive_dates, holiday_dates
            )

            # 字符串形式
            next_term_dates_str = " =X= ".join(
                d.strftime("%Y-%m-%d") for d in next_valid_dates
            )
            next_start_str = next_valid_dates[0].strftime("%Y-%m-%d")
            next_end_str = next_valid_dates[-1].strftime("%Y-%m-%d")

            # 备注：记录「下期」中的节假日冲突日期
            if conflict_dates:
                remark = "節假日衝突：" + " =X= ".join(
                    d.strftime("%Y-%m-%d") for d in conflict_dates
                )

        # 写入一行
        ws.append(
            [
                name,
                next_start_str,
                next_end_str,
                original_date_display,
                lessons if lessons > 0 else "",
                teacher,
                code,
                remark,
                next_term_dates_str,
            ]
        )

    # 设置列宽 & 对齐
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2, max_col=9):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(xlsx_path)
    print(f"✅ 已处理：{os.path.basename(csv_path)} → {os.path.basename(xlsx_path)}")


# ========= 主执行入口 =========
if __name__ == "__main__":
    print("📅 正在更新下期开课日期（使用下期上课日期跳过节假日）...")

    for fname in os.listdir(OUTPUT_DIR):
        if fname.endswith(".csv") and "待更新课程" in fname:
            process_csv(os.path.join(OUTPUT_DIR, fname))

    print("🎉 所有课程日期已更新完毕。")
