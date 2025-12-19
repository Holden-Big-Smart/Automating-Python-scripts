import os
import pandas as pd
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ========= 配置路径 =========
BASE_DIR = os.getcwd()
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

# ========= 固定节假日（不再读取「节假日安排.xlsx」） =========
holiday_dates = {
    datetime(2025, 12, 25).date(),
    datetime(2025, 12, 26).date(),
    datetime(2026, 1, 1).date(),
    datetime(2026, 2, 17).date(),
    datetime(2026, 2, 18).date(),
    datetime(2026, 2, 19).date(),
    datetime(2026, 4, 3).date(),
    datetime(2026, 4, 4).date(),
    datetime(2026, 4, 6).date(),
    datetime(2026, 4, 7).date(),
    datetime(2026, 5, 1).date(),
    datetime(2026, 5, 25).date(),
    datetime(2026, 6, 19).date(),
    datetime(2026, 7, 1).date(),
    datetime(2026, 9, 26).date(),
    datetime(2026, 10, 1).date(),
    datetime(2026, 10, 19).date(),
    datetime(2026, 12, 25).date(),
    datetime(2026, 12, 26).date(),
}

# ========= 工具函数 =========

def extract_lessons(raw):
    match = re.search(r"(\d+)\s*\(堂數\)", str(raw))
    return int(match.group(1)) if match else 0


def adjust_dates_for_holidays(dates, holiday_dates):
    valid_dates = []
    conflict_dates = []

    if not dates:
        return valid_dates, conflict_dates

    # 从原始列表最后一天 +7 开始准备补课日期
    extra_date = dates[-1] + timedelta(days=7)

    for d in dates:
        if d.date() in holiday_dates:
            conflict_dates.append(d)
            while extra_date.date() in holiday_dates:
                extra_date += timedelta(days=7)
            valid_dates.append(extra_date)
            extra_date += timedelta(days=7)
        else:
            valid_dates.append(d)

    return valid_dates, conflict_dates


# ========= 列宽设置（加入新的“逢星期”列） =========
col_widths = {
    "A": 7,   # 逢星期
    "B": 60,
    "C": 15,
    "D": 15,
    "E": 40,
    "F": 8,
    "G": 10,
    "H": 12,
    "I": 25,
    "J": 255,
}


def process_csv(csv_path):
    filename = os.path.basename(csv_path).replace(".csv", "")
    xlsx_path = os.path.join(OUTPUT_DIR, filename + ".xlsx")

    # 读入 CSV → 写入原始数据工作表
    df = pd.read_csv(csv_path, dtype=str).fillna("")
    df.to_excel(xlsx_path, index=False, sheet_name="原始数据")

    wb = load_workbook(xlsx_path)
    ws = wb.create_sheet("日期更新")

    # ========== 新增“逢星期”列，因此所有列后移 ==========
    headers = [
        "逢星期",
        "名稱",
        "下期开课时间",
        "下期结课时间",
        "本期上課日期",
        "堂數",
        "導師",
        "編號",
        "备注",
        "上课日期",
    ]
    ws.append(headers)

    for idx, row in df.iterrows():
        weekday = row.get("逢星期", "")      # 原始数据中 H 列
        name = row.get("名稱", "")
        start_str = str(row.get("上課日期", "")).strip()
        teacher = row.get("導師", "")
        code = row.get("編號", "")
        raw_lesson = row.get("堂數", "")
        lessons = extract_lessons(raw_lesson)

        original_date_display = start_str
        remark = ""
        next_start_str = ""
        next_end_str = ""
        next_term_dates_str = ""

        # 解析「本期开课日」
        try:
            current_start_date = pd.to_datetime(start_str[:10], errors="coerce")
        except Exception:
            current_start_date = None

        if pd.isna(current_start_date) or lessons <= 0:
            remark = "未安排課節"
        else:
            # ========== 1) 生成本期有效上课日期 ==========
            current_naive_dates = [
                current_start_date + timedelta(days=7 * i) for i in range(lessons)
            ]
            current_valid_dates, _ = adjust_dates_for_holidays(current_naive_dates, holiday_dates)

            last_current_date = current_valid_dates[-1]

            # ========== 2) 计算 下期开课时间（保持同星期） ==========
            next_start_date = last_current_date + timedelta(days=7)
            while next_start_date.date() in holiday_dates:
                next_start_date += timedelta(days=7)

            # ========== 3) 生成下期上课日期（节假日顺延） ==========
            next_naive_dates = [
                next_start_date + timedelta(days=7 * i) for i in range(lessons)
            ]
            next_valid_dates, conflict_dates = adjust_dates_for_holidays(
                next_naive_dates, holiday_dates
            )

            next_term_dates_str = "X".join(d.strftime("%Y-%m-%d") for d in next_valid_dates)
            next_start_str = next_valid_dates[0].strftime("%Y-%m-%d")
            next_end_str = next_valid_dates[-1].strftime("%Y-%m-%d")

            if conflict_dates:
                remark = "節假日衝突：" + "X".join(
                    d.strftime("%Y-%m-%d") for d in conflict_dates
                )

        # ========== 写入新行（含新增的“逢星期”列） ==========
        ws.append([
            weekday,
            name,
            next_start_str,
            next_end_str,
            original_date_display,
            lessons if lessons > 0 else "",
            teacher,
            code,
            remark,
            next_term_dates_str,
        ])

    # ========= 设置列宽 & 对齐 =========
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2, max_col=10):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(xlsx_path)
    print(f"✅ 已处理：{os.path.basename(csv_path)} → {os.path.basename(xlsx_path)}")


# ========= 主执行入口 =========
if __name__ == "__main__":
    print("📅 正在更新下期开课日期（含逢星期列）...")

    for fname in os.listdir(OUTPUT_DIR):
        if fname.endswith(".csv") and "待更新课程" in fname:
            process_csv(os.path.join(OUTPUT_DIR, fname))

    print("🎉 所有课程日期已更新完毕。")
