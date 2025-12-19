import os
import pandas as pd
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ========= 配置路径 =========
BASE_DIR = os.getcwd()
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
HOLIDAY_PATH = os.path.join(BASE_DIR, "节假日安排.xlsx")

# ========= 加载节假日列表 =========
holidays = pd.read_excel(HOLIDAY_PATH, dtype=str)
holiday_dates = set(pd.to_datetime(holidays.iloc[:, 0], errors="coerce").dropna().dt.date)

# ========= 工具函数 =========

def extract_lessons(raw):
    match = re.search(r"(\d+)\s*\(堂數\)", str(raw))
    return int(match.group(1)) if match else 0

def skip_holidays(start_date):
    while start_date in holiday_dates:
        start_date += timedelta(days=1)
    return start_date

col_widths = {
    "A": 18, "B": 10, "C": 10, "D": 30,
    "E": 10, "F": 10, "G": 10, "H": 20
}

def process_csv(csv_path):
    filename = os.path.basename(csv_path).replace(".csv", "")
    xlsx_path = os.path.join(OUTPUT_DIR, filename + ".xlsx")

    df = pd.read_csv(csv_path, dtype=str).fillna("")
    df.to_excel(xlsx_path, index=False, sheet_name="原始数据")

    wb = load_workbook(xlsx_path)
    ws = wb.create_sheet("日期更新")

    headers = ["名稱", "下期开课时间", "下期结课时间", "本期上課日期", "堂數", "導師", "編號", "备注"]
    ws.append(headers)

    for _, row in df.iterrows():
        name = row.get("名稱", "")
        start_str = str(row.get("上課日期", "")).strip()
        teacher = row.get("導師", "")
        code = row.get("編號", "")
        raw_lesson = row.get("堂數", "")
        lessons = extract_lessons(raw_lesson)
        remark = ""

        # 使用原始字符串写入D列
        original_date_display = start_str

        # 提取开始日期作为计算基准
        try:
            start_date = pd.to_datetime(start_str[:10], errors="coerce")
        except:
            start_date = None
            remark = "無法解析上課日期"

        if pd.isna(start_date) or lessons <= 0:
            remark = "未安排課節"
            next_start, next_end = None, None
        else:
            try:
                end_date = start_date + timedelta(weeks=lessons - 1)
                next_start = skip_holidays(end_date + timedelta(days=7))
                raw_next_end = next_start + timedelta(weeks=lessons - 1)
                next_end = skip_holidays(raw_next_end)
            except:
                next_start, next_end = None, None
                remark = "日期計算錯誤"

        # 写入内容
        ws.append([
            name,
            next_start.strftime("%Y-%m-%d") if next_start else "",
            next_end.strftime("%Y-%m-%d") if next_end else "",
            original_date_display,
            lessons if lessons > 0 else "",
            teacher,
            code,
            remark
        ])

    # 设置列宽 & 左对齐
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(xlsx_path)
    print(f"✅ 已处理：{os.path.basename(csv_path)} → {os.path.basename(xlsx_path)}")

# ========= 主执行入口 =========
if __name__ == "__main__":
    print("📅 正在更新下期开课日期...")

    for fname in os.listdir(OUTPUT_DIR):
        if fname.endswith(".csv") and "待更新课程" in fname:
            process_csv(os.path.join(OUTPUT_DIR, fname))

    print("🎉 所有课程日期已更新完毕。")
