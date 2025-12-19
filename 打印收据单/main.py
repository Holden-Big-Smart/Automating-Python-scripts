import pyautogui
import openpyxl
import os
import time
import keyboard
import threading
import cv2
import numpy as np

# ============================================================
# 模拟真实鼠标移动（速度更快，更自然）
# ============================================================
def move_mouse_realistically(x, y, speed=1000):
    current_x, current_y = pyautogui.position()
    distance = ((x - current_x) ** 2 + (y - current_y) ** 2) ** 0.5
    duration = distance / speed
    pyautogui.moveTo(x, y, duration=duration)

# ============================================================
# 封装函数：尝试识别
# ============================================================
def 尝试识别(图像路径, 是否点击=1, 尝试间隔=1, 识别精度=0.9, 区域=[0, 0, 1365, 767]):
    print(f"[尝试识别] 当前图片：{图像路径}")
    while True:
        screenshot = pyautogui.screenshot(region=tuple(区域))
        screenshot = cv2.cvtColor(np.array(screenshot), cv2.COLOR_RGB2BGR)
        template = cv2.imread(图像路径, cv2.IMREAD_COLOR)

        result = cv2.matchTemplate(screenshot, template, cv2.TM_CCOEFF_NORMED)
        _, max_val, _, max_loc = cv2.minMaxLoc(result)

        if max_val >= 识别精度:
            h, w = template.shape[:2]
            x = 区域[0] + max_loc[0] + w // 2
            y = 区域[1] + max_loc[1] + h // 2
            move_mouse_realistically(x, y)
            if 是否点击:
                pyautogui.click()
            return (x, y)
        else:
            time.sleep(尝试间隔)

# ============================================================
# 全局紧急停止线程：数字小键盘 0 立即强退
# ============================================================
def emergency_kill_listener():
    while True:
        if keyboard.is_pressed("num 0"):
            print("\n[紧急中止] 检测到 Num 0，强制退出脚本")
            os._exit(0)

threading.Thread(target=emergency_kill_listener, daemon=True).start()

# ============================================================
# 主流程逻辑
# ============================================================
def main():
    print("📢 脚本即将在5秒后执行，请转至指定界面以便识别")
    time.sleep(1)

    待打印路径 = "待打印数据.xlsx"
    已完成路径 = "已完成数据.xlsx"

    while True:
        wb = openpyxl.load_workbook(待打印路径)
        ws = wb["课程收据"]
        member_id = ws["B2"].value

        if not member_id:
            print("✅ 所有数据处理完毕，退出脚本")
            break

        print(f"\n🆔 正在处理：{member_id}")

        # 阶段1：搜索输入
        x, y = 尝试识别("assets/收据编号.png", 0)
        pyautogui.moveTo(x, y + 25)
        time.sleep(0.5)
        pyautogui.click()
        time.sleep(0.5)
        pyautogui.hotkey("ctrl", "a")
        time.sleep(0.5)
        pyautogui.typewrite(str(member_id))
        time.sleep(0.5)
        pyautogui.press("enter")

        # 阶段2：打印流程
        尝试识别("assets/列印.png", 1)
        尝试识别("assets/A4单收据.png", 1)
        尝试识别("assets/页面-全部.png", 1)
        尝试识别("assets/仅限奇数页.png", 1)
        尝试识别("assets/检查页面_仅限奇数页.png", 0)
        尝试识别("assets/确认列印.png", 0)

        # 数据移动
        completed_wb = openpyxl.load_workbook(已完成路径)
        completed_ws = completed_wb.active
        source_row = ws["B2"].row
        values = [cell.value for cell in ws[source_row]]
        completed_ws.append(values)
        ws.delete_rows(source_row)
        wb.save(待打印路径)
        completed_wb.save(已完成路径)

        print(f"✅ 已处理并转移：{member_id}")
        time.sleep(1)

# ============================================================
# 程序入口
# ============================================================
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ 脚本错误：{e}")
