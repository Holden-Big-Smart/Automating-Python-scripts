import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

# ========= 日期解析：输入格式 日.月-日.月.年 =========
def parse_date_input(user_input: str):
    s = user_input.strip().replace('/', '.')
    if '-' not in s or s.count('.') < 3:
        raise ValueError("输入格式应为：日.月-日.月.年 例如：30.10-9.11.2025 或 07.11-03.01.2026")
    left, right = s.split('-', 1)
    right_parts = right.split('.')
    if len(right_parts) != 3:
        raise ValueError("右半段必须包含 年")
    def _to_int(x): return int(x.strip())
    sd_str, sm_str = left.split('.')
    ed_str, em_str, y_str = right_parts
    sd, sm, ed, em, end_year = _to_int(sd_str), _to_int(sm_str), _to_int(ed_str), _to_int(em_str), _to_int(y_str)
    start_year = end_year - 1 if em < sm else end_year
    start_date = datetime(start_year, sm, sd)
    end_date = datetime(end_year, em, ed)
    excel_var = f"{sd}/{sm}/{start_year}-{ed}/{em}/{end_year}"
    display = f"{sd}.{sm}-{ed}.{em}.{end_year}"
    filename = f"{display}.xlsx"
    return start_date, end_date, excel_var, display, filename


# ========= 主流程 =========
def main():
    user_input = input("请输入数据时间（格式：日.月-日.月.年，例如 30.10-9.11.2025 或 07.11-03.01.2026）：").strip()
    start_date, end_date, excel_var, display_str, filename_str = parse_date_input(user_input)

    print(f"数据日期（Excel变量）：{excel_var}")
    print(f"Excel显示日期：{display_str}")
    print(f"输出文件名：{filename_str}")

    # ===== 課程收據 =====
    df_course = pd.read_csv("屯門婦聯 - 會員及課程管理系統 - 課程收據.csv")
    drop_cols_course = ["服務中心","報名中心","會員編號","會員姓名(英文)","會員類別","課程收費","其他收費","扣減","日期","撤銷"]
    df_course = df_course.drop(columns=[c for c in drop_cols_course if c in df_course.columns], errors="ignore")

    if "經辨人員" in df_course.columns:
        if "編號" in df_course.columns:
            df_course = df_course.sort_values(by=["經辨人員", "編號"], ascending=[False, True])
        else:
            df_course = df_course.sort_values(by="經辨人員", ascending=False)

    if "#" in df_course.columns:
        n = len(df_course)
        if n > 1:
            col_idx = df_course.columns.get_loc("#")
            df_course.iloc[:-1, col_idx] = range(1, n)
        else:
            df_course["#"] = [1]

    df_course["總額"] = df_course["總額"].replace(r'[\$,]', '', regex=True).astype(float)
    課程總額 = df_course["總額"].iloc[-1] if not df_course["總額"].empty else 0.0

    if "經辨人員" in df_course.columns:
        df_course_valid = df_course[
            df_course["經辨人員"].notna() &
            (df_course["經辨人員"].astype(str).str.strip() != "")
        ]
    else:
        df_course_valid = df_course.copy()

    group_course = df_course_valid.groupby("付款方式")["總額"].sum().to_dict()

    # ===== 會費收據 =====
    df_fee = pd.read_csv("屯門婦聯 - 會員及課程管理系統 - 會費收據.csv")
    drop_cols_fee = ["中心","會員編號","會員姓名(英文)","日期","撤銷"]
    df_fee = df_fee.drop(columns=[c for c in drop_cols_fee if c in df_fee.columns], errors="ignore")

    if "經辨人員" in df_fee.columns:
        if "編號" in df_fee.columns:
            df_fee = df_fee.sort_values(by=["經辨人員", "編號"], ascending=[False, True])
        else:
            df_fee = df_fee.sort_values(by="經辨人員", ascending=False)

    if "#" in df_fee.columns:
        n_fee = len(df_fee)
        if n_fee > 1:
            col_idx_fee = df_fee.columns.get_loc("#")
            df_fee.iloc[:-1, col_idx_fee] = range(1, n_fee)
        else:
            df_fee["#"] = [1]

    df_fee["會費"] = df_fee["會費"].replace(r'[\$,]', '', regex=True).astype(float)
    會費總額 = df_fee["會費"].iloc[-1] if not df_fee["會費"].empty else 0.0

    if "經辨人員" in df_fee.columns:
        df_fee_valid = df_fee[
            df_fee["經辨人員"].notna() &
            (df_fee["經辨人員"].astype(str).str.strip() != "")
        ]
    else:
        df_fee_valid = df_fee.copy()

    group_fee = df_fee_valid.groupby("付款方式")["會費"].sum().to_dict()

    # ===== 合併總額及付款方式 =====
    最終總額 = 課程總額 + 會費總額
    all_methods = set(group_course.keys()) | set(group_fee.keys())
    merged_method_totals = {m: group_course.get(m, 0) + group_fee.get(m, 0) for m in all_methods}

    merged_line1 = f"{excel_var} 【總計】${最終總額:.0f}"
    merged_line2 = "  ".join([f"{m}${merged_method_totals[m]:.0f}" for m in merged_method_totals])
    total_summary = merged_line1 + "\n" + merged_line2

    # ===== 各經辨人員統計 =====
    merged_vars = {}
    all_staff = sorted(
        set(df_course_valid["經辨人員"].unique()) |
        set(df_fee_valid["經辨人員"].unique())
    )

    for staff in all_staff:
        if pd.isna(staff) or str(staff).strip() == "":
            continue
        course_total = df_course_valid[df_course_valid["經辨人員"] == staff]["總額"].sum()
        fee_total = df_fee_valid[df_fee_valid["經辨人員"] == staff]["會費"].sum()
        staff_total = course_total + fee_total
        course_group_staff = df_course_valid[df_course_valid["經辨人員"] == staff].groupby("付款方式")["總額"].sum().to_dict()
        fee_group_staff = df_fee_valid[df_fee_valid["經辨人員"] == staff].groupby("付款方式")["會費"].sum().to_dict()
        all_staff_methods = set(course_group_staff.keys()) | set(fee_group_staff.keys())
        merged_staff_methods = {m: course_group_staff.get(m, 0) + fee_group_staff.get(m, 0) for m in all_staff_methods}

        line1 = f"{staff}【總計】${staff_total:.0f}"
        line2 = "  ".join([f"{m}${merged_staff_methods[m]:.0f}" for m in merged_staff_methods])
        merged_vars[staff] = line1 + "\n" + line2

    # ===== 導出Excel =====
    with pd.ExcelWriter(filename_str, engine="openpyxl") as writer:
        df_course.to_excel(writer, sheet_name="課程收據", index=False)
        df_fee.to_excel(writer, sheet_name="會費收據", index=False)
        pd.DataFrame(columns=["統計"]).to_excel(writer, sheet_name="統計結果", index=False)

        wb = writer.book
        ws = wb["統計結果"]

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font_title = Font(size=14)
        font_body = Font(size=12)
        row_height_14 = 20
        row_height_12 = 16

        # 写入总汇
        ws["A1"] = total_summary
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        ws["A1"].font = font_title
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = row_height_14 * (total_summary.count("\n") + 1)

        # 写入各经辨人員
        row_idx = 2
        for staff, text in merged_vars.items():
            ws.cell(row=row_idx, column=1, value=text)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
            cell = ws.cell(row=row_idx, column=1)
            cell.font = font_body
            cell.alignment = align_center
            ws.row_dimensions[row_idx].height = row_height_12 * (text.count("\n") + 1)
            row_idx += 1

        ws.column_dimensions[get_column_letter(1)].width = 60

        ws2 = wb.create_sheet("數據信息")
        ws2["A1"] = "數據日期"; ws2["B1"] = excel_var
        ws2["A2"] = "顯示日期"; ws2["B2"] = display_str
        ws2["A3"] = "開始日期(ISO)"; ws2["B3"] = start_date.strftime("%Y-%m-%d")
        ws2["A4"] = "結束日期(ISO)"; ws2["B4"] = end_date.strftime("%Y-%m-%d")

    print(f"\n✅ 已输出文件：{filename_str}")


if __name__ == "__main__":
    main()
