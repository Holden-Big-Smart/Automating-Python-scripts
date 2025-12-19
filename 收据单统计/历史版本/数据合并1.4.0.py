import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl import Workbook
from copy import copy as shallow_copy
import traceback
import sys
import os

# ========== 控制台提示/退出 ==========
def pause_and_exit(code=0):
    try:
        input("\n按回车键退出 ...")
    except Exception:
        pass
    sys.exit(code)

def error(msg: str):
    print(f"【错误】{msg}")

def warn(msg: str):
    print(f"【警告】{msg}")

def info(msg: str):
    print(f"【信息】{msg}")

# ========== 关键列定义（含别名）==========
# 注意：逻辑字段必须解析成功，否则报错并暂停
REQUIRED_COURSE_FIELDS = {
    "#":            ["#", "序號", "序号"],
    "編號":          ["編號", "编号", "單號", "单号", "收據編號", "收据编号"],
    "會員姓名(中文)": ["會員姓名(中文)", "会员姓名(中文)", "會員姓名（中文）", "会员姓名（中文）", "會員姓名", "会员姓名", "姓名"],
    "課程名稱":       ["課程名稱", "课程名称", "課程名", "课程名", "課程", "课程"],
    "總額":          ["總額", "总额", "金額", "金额", "收款金額", "收款金额"],
    "付款方式":       ["付款方式", "支付方式"],
    "經辨人員":       ["經辨人員", "經辦人員", "经辨人员", "经办人员", "經辦", "经办"],
}

REQUIRED_FEE_FIELDS = {
    "#":            ["#", "序號", "序号"],
    "編號":          ["編號", "编号", "單號", "单号", "收據編號", "收据编号"],
    "會員姓名(中文)": ["會員姓名(中文)", "会员姓名(中文)", "會員姓名（中文）", "会员姓名（中文）", "會員姓名", "会员姓名", "姓名"],
    "會員類別":       ["會員類別", "会员类别", "會員類型", "会员类型", "類別", "类别"],
    "會費":          ["會費", "会费", "金額", "金额", "收款金額", "收款金额"],
    "付款方式":       ["付款方式", "支付方式"],
    "經辨人員":       ["經辨人員", "經辦人員", "经辨人员", "经办人员", "經辦", "经办"],
}

def resolve_required_columns(df: pd.DataFrame, field_alias_map: dict, table_alias: str) -> dict:
    """
    将逻辑字段映射为表内真实列名；任一解析失败即报错并暂停。
    返回：{逻辑字段: 实际列名}
    """
    resolved = {}
    missing = []
    for logical, candidates in field_alias_map.items():
        hit = None
        for c in candidates:
            if c in df.columns:
                hit = c
                break
        if hit is None:
            missing.append(f"{logical}（候选：{', '.join(candidates)}）")
        else:
            resolved[logical] = hit

    if missing:
        error(f"{table_alias} 缺少关键列：{'; '.join(missing)}。请检查源CSV。")
        pause_and_exit(1)
    return resolved

# ========= 日期解析：输入格式 日.月-日.月.年 =========
def parse_date_input(user_input: str):
    s = user_input.strip().replace("/", ".")
    if "-" not in s or s.count(".") < 3:
        raise ValueError("输入格式应为：日.月-日.月.年 例如：30.10-9.11.2025 或 07.11-03.01.2026")
    left, right = s.split("-", 1)
    right_parts = right.split(".")
    if len(right_parts) != 3:
        raise ValueError("右半段必须包含 年")

    def _to_int(x): return int(x.strip())

    sd_str, sm_str = left.split(".")
    ed_str, em_str, y_str = right_parts
    sd, sm, ed, em, end_year = _to_int(sd_str), _to_int(sm_str), _to_int(ed_str), _to_int(em_str), _to_int(y_str)
    start_year = end_year - 1 if em < sm else end_year
    start_date = datetime(start_year, sm, sd)
    end_date = datetime(end_year, em, ed)
    excel_var = f"{sd}/{sm}/{start_year}-{ed}/{em}/{end_year}"
    display   = f"{sd}.{sm}-{ed}.{em}.{end_year}"
    filename  = f"{display}.xlsx"
    return start_date, end_date, excel_var, display, filename

# ========= 安全复制单元格样式，规避 StyleProxy 报错 =========
def safe_copy_cell_style(src_cell, tgt_cell):
    try:
        if src_cell.has_style:
            # Font
            try:
                tgt_cell.font = shallow_copy(src_cell.font)
            except Exception:
                f = src_cell.font
                try:
                    tgt_cell.font = Font(
                        name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                        vertAlign=f.vertAlign, underline=f.underline, strike=f.strike, color=f.color
                    )
                except Exception:
                    pass

            # Fill
            try:
                tgt_cell.fill = shallow_copy(src_cell.fill)
            except Exception:
                pass

            # Border
            try:
                tgt_cell.border = shallow_copy(src_cell.border)
            except Exception:
                pass

            # Alignment
            try:
                tgt_cell.alignment = shallow_copy(src_cell.alignment)
            except Exception:
                a = src_cell.alignment
                try:
                    tgt_cell.alignment = Alignment(
                        horizontal=a.horizontal,
                        vertical=a.vertical,
                        textRotation=a.textRotation,
                        wrap_text=(a.wrapText if hasattr(a, "wrapText") else getattr(a, "wrap_text", None)),
                        shrink_to_fit=(a.shrinkToFit if hasattr(a, "shrinkToFit") else getattr(a, "shrink_to_fit", None)),
                        indent=a.indent,
                    )
                except Exception:
                    pass

            # Number format
            try:
                tgt_cell.number_format = src_cell.number_format
            except Exception:
                pass

            # Protection
            try:
                tgt_cell.protection = shallow_copy(src_cell.protection)
            except Exception:
                pass
    except Exception:
        pass

# ========= 将一个 sheet 的内容（值+样式）复制到目标 sheet =========
def copy_sheet_content(src_ws, tgt_ws, start_row: int, preserve_merges: bool = True, blank_row_after: bool = True):
    """
    把 src_ws 的所有单元格（值与样式）复制到 tgt_ws，从 start_row 开始写入。
    - preserve_merges=True 时，复制源表的合并单元格到目标（偏移后）。
    - preserve_merges=False（用于“統計結果”），只复制值与样式，不复制合并。
    - 表之间插入空行：blank_row_after=True。
    返回：下一次可写的起始行号。
    """
    max_row = src_ws.max_row
    max_col = src_ws.max_column

    # 复制列宽（只复制到源表的 max_col）
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        src_dim = src_ws.column_dimensions.get(letter)
        if src_dim and src_dim.width:
            tgt_ws.column_dimensions[letter].width = src_dim.width

    # 复制行高与单元格（值+样式）
    for r in range(1, max_row + 1):
        tgt_r = start_row + r - 1

        # 行高
        rd = src_ws.row_dimensions.get(r)
        if rd and rd.height:
            tgt_ws.row_dimensions[tgt_r].height = rd.height

        # 单元格
        for c in range(1, max_col + 1):
            src_cell = src_ws.cell(row=r, column=c)
            tgt_cell = tgt_ws.cell(row=tgt_r, column=c, value=src_cell.value)
            safe_copy_cell_style(src_cell, tgt_cell)

    # 复制合并单元格（偏移后）
    if preserve_merges and src_ws.merged_cells and len(src_ws.merged_cells.ranges) > 0:
        for m_range in src_ws.merged_cells.ranges:
            min_col = m_range.min_col
            max_col_merge = m_range.max_col
            min_row = m_range.min_row + (start_row - 1)
            max_row_merge = m_range.max_row + (start_row - 1)
            tgt_ws.merge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row_merge,
                end_column=max_col_merge
            )

    # 计算下一起始行（表间空一行）
    next_start_row = start_row + max_row + (1 if blank_row_after else 0)
    return next_start_row

# ========= 安全读取 CSV =========
def read_csv_safe(path: str) -> pd.DataFrame:
    try:
        if not os.path.exists(path):
            error(f"读取失败：文件不存在 -> {path}")
            pause_and_exit(1)
        df = pd.read_csv(path)
        if df.empty:
            warn(f"文件为空：{path}")
        return df
    except FileNotFoundError:
        error(f"读取失败：找不到文件 -> {path}")
        pause_and_exit(1)
    except PermissionError:
        error(f"读取失败：没有权限访问 -> {path}")
        pause_and_exit(1)
    except UnicodeDecodeError:
        error(f"读取失败：编码错误，请确认 CSV 编码格式 -> {path}")
        pause_and_exit(1)
    except Exception as e:
        error(f"读取失败：{path} -> {e}")
        traceback.print_exc()
        pause_and_exit(1)

# ========= 将金额列安全转为数值 =========
def to_float_safe(df: pd.DataFrame, col: str, table_alias: str):
    if col not in df.columns:
        error(f"{table_alias} 缺少关键金额列「{col}」，无法继续。")
        pause_and_exit(1)
    try:
        df[col] = df[col].replace(r"[\$,]", "", regex=True)
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0).astype(float)
        return df
    except Exception as e:
        error(f"{table_alias} 金额列「{col}」清洗失败：{e}")
        traceback.print_exc()
        pause_and_exit(1)

# ========= 主流程 =========
def main():
    try:
        user_input = input("请输入数据时间（格式：日.月-日.月.年，例如 30.10-9.11.2025 或 07.11-03.01.2026）：").strip()
        start_date, end_date, excel_var, display_str, filename_str = parse_date_input(user_input)

        print(f"数据日期（Excel变量）：{excel_var}")
        print(f"Excel显示日期：{display_str}")
        print(f"输出文件名：{filename_str}")

        # ===== 读取 CSV =====
        df_course = read_csv_safe("屯門婦聯 - 會員及課程管理系統 - 課程收據.csv")
        df_fee    = read_csv_safe("屯門婦聯 - 會員及課程管理系統 - 會費收據.csv")

        # ===== 强制解析关键列（含别名）=====
        colmap_course = resolve_required_columns(df_course, REQUIRED_COURSE_FIELDS, "課程收據")
        colmap_fee    = resolve_required_columns(df_fee,    REQUIRED_FEE_FIELDS,    "會費收據")

        # ===== 清理非必需列（保持你原有的列删除逻辑，不影响关键列）=====
        drop_cols_course = ["服務中心","報名中心","會員編號","會員姓名(英文)","會員類別","課程收費","其他收費","扣減","日期","撤銷"]
        df_course = df_course.drop(columns=[c for c in drop_cols_course if c in df_course.columns], errors="ignore")

        drop_cols_fee = ["中心","會員編號","會員姓名(英文)","日期","撤銷"]
        df_fee = df_fee.drop(columns=[c for c in drop_cols_fee if c in df_fee.columns], errors="ignore")

        # ===== 課程收據：排序（经办人 + 編號）=====
        if colmap_course["編號"] in df_course.columns:
            df_course = df_course.sort_values(by=[colmap_course["經辨人員"], colmap_course["編號"]], ascending=[False, True])
        else:
            df_course = df_course.sort_values(by=colmap_course["經辨人員"], ascending=False)

        # 序号列（覆盖到倒数第二行）
        if colmap_course["#"] in df_course.columns:
            n = len(df_course)
            if n > 1:
                col_idx = df_course.columns.get_loc(colmap_course["#"])
                df_course.iloc[:-1, col_idx] = range(1, n)
            else:
                df_course[colmap_course["#"]] = [1]

        # 金额清洗
        df_course = to_float_safe(df_course, colmap_course["總額"], "課程收據")
        課程總額 = df_course[colmap_course["總額"]].iloc[-1] if not df_course[colmap_course["總額"]].empty else 0.0

        # 有效行（经办人非空）
        df_course_valid = df_course[df_course[colmap_course["經辨人員"]].astype(str).str.strip() != ""]

        # 付款方式汇总
        try:
            group_course = df_course_valid.groupby(colmap_course["付款方式"])[colmap_course["總額"]].sum().to_dict()
        except Exception as e:
            error(f"課程收據按『{colmap_course['付款方式']}』汇总失败：{e}")
            traceback.print_exc()
            pause_and_exit(1)

        # ===== 會費收據：排序（经办人 + 編號）=====
        if colmap_fee["編號"] in df_fee.columns:
            df_fee = df_fee.sort_values(by=[colmap_fee["經辨人員"], colmap_fee["編號"]], ascending=[False, True])
        else:
            df_fee = df_fee.sort_values(by=colmap_fee["經辨人員"], ascending=False)

        # 序号列
        if colmap_fee["#"] in df_fee.columns:
            n_fee = len(df_fee)
            if n_fee > 1:
                col_idx_fee = df_fee.columns.get_loc(colmap_fee["#"])
                df_fee.iloc[:-1, col_idx_fee] = range(1, n_fee)
            else:
                df_fee[colmap_fee["#"]] = [1]

        # 金额清洗
        df_fee = to_float_safe(df_fee, colmap_fee["會費"], "會費收據")
        會費總額 = df_fee[colmap_fee["會費"]].iloc[-1] if not df_fee[colmap_fee["會費"]].empty else 0.0

        df_fee_valid = df_fee[df_fee[colmap_fee["經辨人員"]].astype(str).str.strip() != ""]

        try:
            group_fee = df_fee_valid.groupby(colmap_fee["付款方式"])[colmap_fee["會費"]].sum().to_dict()
        except Exception as e:
            error(f"會費收據按『{colmap_fee['付款方式']}』汇总失败：{e}")
            traceback.print_exc()
            pause_and_exit(1)

        # ===== 合併總額及付款方式 =====
        最終總額 = float(課程總額) + float(會費總額)
        all_methods = set(group_course.keys()) | set(group_fee.keys())
        merged_method_totals = {m: group_course.get(m, 0) + group_fee.get(m, 0) for m in all_methods}

        merged_line1 = f"{excel_var} 【總計】${最終總額:.0f}"
        merged_line2 = "  ".join([f"{m}${merged_method_totals[m]:.0f}" for m in merged_method_totals])
        total_summary = merged_line1 + ("\n" + merged_line2 if merged_line2 else "")

        # ===== 各經辨人員統計 =====
        staff_set = set()
        staff_set |= set(df_course_valid[colmap_course["經辨人員"]].dropna().astype(str).str.strip().unique())
        staff_set |= set(df_fee_valid[colmap_fee["經辨人員"]].dropna().astype(str).str.strip().unique())
        all_staff = sorted([s for s in staff_set if s])

        merged_vars = {}
        for staff in all_staff:
            # 各来源筛选
            course_total = df_course_valid[df_course_valid[colmap_course["經辨人員"]].astype(str).str.strip() == staff][colmap_course["總額"]].sum()
            fee_total    = df_fee_valid[df_fee_valid[colmap_fee["經辨人員"]].astype(str).str.strip() == staff][colmap_fee["會費"]].sum()
            staff_total  = float(course_total) + float(fee_total)

            # 各方式
            try:
                course_group_staff = (
                    df_course_valid[df_course_valid[colmap_course["經辨人員"]].astype(str).str.strip() == staff]
                    .groupby(colmap_course["付款方式"])[colmap_course["總額"]].sum().to_dict()
                )
            except Exception as e:
                error(f"課程收據员『{staff}』方式汇总失败：{e}")
                traceback.print_exc()
                pause_and_exit(1)

            try:
                fee_group_staff = (
                    df_fee_valid[df_fee_valid[colmap_fee["經辨人員"]].astype(str).str.strip() == staff]
                    .groupby(colmap_fee["付款方式"])[colmap_fee["會費"]].sum().to_dict()
                )
            except Exception as e:
                error(f"會費收據员『{staff}』方式汇总失败：{e}")
                traceback.print_exc()
                pause_and_exit(1)

            all_staff_methods = set(course_group_staff.keys()) | set(fee_group_staff.keys())
            merged_staff_methods = {m: course_group_staff.get(m, 0) + fee_group_staff.get(m, 0) for m in all_staff_methods}

            line1 = f"{staff}【總計】${staff_total:.0f}"
            line2 = "  ".join([f"{m}${merged_staff_methods[m]:.0f}" for m in merged_staff_methods])
            merged_vars[staff] = line1 + ("\n" + line2 if line2 else "")

        # ===== 导出Excel（含“最终输出”拼接）=====
        with pd.ExcelWriter(filename_str, engine="openpyxl") as writer:
            # 先把三张数据表写出去
            df_course.to_excel(writer, sheet_name="課程收據", index=False)
            df_fee.to_excel(writer, sheet_name="會費收據", index=False)
            pd.DataFrame(columns=["統計"]).to_excel(writer, sheet_name="統計結果", index=False)

            wb = writer.book

            # === 填充“統計結果”的展示（1×7 合并行）===
            ws_stat = wb["統計結果"]
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            font_title = Font(size=14, bold=True)  # 第一行标题加粗
            font_body  = Font(size=12)
            row_height_14 = 20
            row_height_12 = 16

            # 总汇
            ws_stat["A1"] = total_summary
            ws_stat.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
            ws_stat["A1"].font = font_title
            ws_stat["A1"].alignment = align_center
            ws_stat.row_dimensions[1].height = row_height_14 * (total_summary.count("\n") + 1 if total_summary else 1)

            # 各经辨人員
            row_idx = 2
            for staff, text in merged_vars.items():
                ws_stat.cell(row=row_idx, column=1, value=text)
                ws_stat.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
                cell = ws_stat.cell(row=row_idx, column=1)
                cell.font = font_body
                cell.alignment = align_center
                ws_stat.row_dimensions[row_idx].height = row_height_12 * (text.count("\n") + 1 if text else 1)
                row_idx += 1

            ws_stat.column_dimensions[get_column_letter(1)].width = 60

            # 数据信息
            ws_info = wb.create_sheet("數據信息")
            ws_info["A1"] = "數據日期";       ws_info["B1"] = excel_var
            ws_info["A2"] = "顯示日期";       ws_info["B2"] = display_str
            ws_info["A3"] = "開始日期(ISO)"; ws_info["B3"] = start_date.strftime("%Y-%m-%d")
            ws_info["A4"] = "結束日期(ISO)"; ws_info["B4"] = end_date.strftime("%Y-%m-%d")

            # === 生成“最终输出”，并把三张表按次序拼接（保留样式）===
            if "最终输出" in wb.sheetnames:
                del wb["最终输出"]
            ws_out = wb.create_sheet("最终输出")

            current_row = 1

            # 1) 課程收據：保留合并
            if "課程收據" in wb.sheetnames:
                current_row = copy_sheet_content(wb["課程收據"], ws_out, current_row, preserve_merges=True, blank_row_after=True)

            # 2) 會費收據：保留合并
            if "會費收據" in wb.sheetnames:
                current_row = copy_sheet_content(wb["會費收據"], ws_out, current_row, preserve_merges=True, blank_row_after=True)

            # 3) 統計結果：先不带合并复制，随后逐行执行 1~7 列合并
            if "統計結果" in wb.sheetnames:
                start_row_stat_in_out = current_row
                current_row = copy_sheet_content(wb["統計結果"], ws_out, current_row, preserve_merges=False, blank_row_after=True)

                # 逐行合并 1~7 列（仅对有值的行）
                max_row_copied = current_row - 1
                for r in range(start_row_stat_in_out, max_row_copied + 1):
                    val = ws_out.cell(row=r, column=1).value
                    if val is not None and str(val).strip() != "":
                        ws_out.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

            # === 设置“最终输出”列宽 ===
            ws_out.column_dimensions["A"].width = 2.15
            ws_out.column_dimensions["B"].width = 12
            ws_out.column_dimensions["C"].width = 8
            ws_out.column_dimensions["D"].width = 32
            ws_out.column_dimensions["E"].width = 7
            ws_out.column_dimensions["F"].width = 8.46
            ws_out.column_dimensions["G"].width = 8.54

            # === 重设“最终输出”C列为左对齐（尽量保留原有的其它对齐属性）===
            max_r = ws_out.max_row
            for r in range(1, max_r + 1):
                cell = ws_out.cell(row=r, column=3)  # C列
                a = cell.alignment if cell.has_style else Alignment()
                try:
                    new_align = Alignment(
                        horizontal="left",
                        vertical=a.vertical,
                        textRotation=a.textRotation,
                        wrap_text=(a.wrapText if hasattr(a, "wrapText") else getattr(a, "wrap_text", None)),
                        shrink_to_fit=(a.shrinkToFit if hasattr(a, "shrinkToFit") else getattr(a, "shrink_to_fit", None)),
                        indent=a.indent,
                    )
                except Exception:
                    new_align = Alignment(horizontal="left")
                cell.alignment = new_align

            # === 将“最终输出”移动到最前 ===
            sheets = wb._sheets
            out_sheet = wb["最终输出"]
            sheets.remove(out_sheet)
            sheets.insert(0, out_sheet)

        print(f"\n✅ 已输出文件：{filename_str}")
        print("✅ 已生成“最终输出”工作表，并置于最前；三个来源工作表按顺序完整拼接，表间空一行；"
              "已按要求设置列宽（A:2.15, B:12, C:8, D:32, E:7, F:8.46, G:8.54），且将 C 列统一左对齐。")
        pause_and_exit(0)

    except ValueError as ve:
        error(f"输入/日期解析异常：{ve}")
        traceback.print_exc()
        pause_and_exit(1)
    except KeyboardInterrupt:
        warn("用户中断。")
        pause_and_exit(1)
    except Exception as e:
        error(f"未处理异常：{e}")
        traceback.print_exc()
        pause_and_exit(1)

if __name__ == "__main__":
    main()
